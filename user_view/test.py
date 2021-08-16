import mysql.connector

from openpyxl import load_workbook, Workbook
import openpyxl as px

import subprocess
import os
import sys

def scrape_sheet(fil):
    print("scrp-sheet")
    wb_obj = px.load_workbook(fil)
    sheet_obj = wb_obj.active
      
    max_col = sheet_obj.max_column
    
    field=[]
    val1=[]
    val2=[]
    val3=[]
    quality=[]  
    composition=[]
    
    
    for i in range(1, max_col + 1):
        cell_obj = sheet_obj.cell(row = i, column = 1)
        if cell_obj.value == 'FRINGE':
            # print('fringe found-'+str(i))
            field.append('Fringe')
            val1.append(sheet_obj.cell(i,2).value)
            val2.append(sheet_obj.cell(i,3).value)
            val3.append(sheet_obj.cell(i,4).value)
        if cell_obj.value == 'MEND':
            # print('mend found-'+str(i))
            field.append('Mend')
            val1.append(sheet_obj.cell(i,2).value)
            val2.append(sheet_obj.cell(i,3).value)
            val3.append(sheet_obj.cell(i,4).value)
        if cell_obj.value == 'GR. PERCH':
            # print('perch found-'+str(i))
            field.append('Greasy Perch')
            val1.append(sheet_obj.cell(i,2).value)
            val2.append(sheet_obj.cell(i,3).value)
            val3.append(sheet_obj.cell(i,4).value)
            
            
        cell_obj2 = sheet_obj.cell(row = i, column = 6)
        if cell_obj2.value == 'MACHINE No:':
            # print('MACHINE No: found-'+str(i))
            field.append('Machine No.')
            val1.append(sheet_obj.cell(i,7).value)
            val2.append(sheet_obj.cell(i,7).value)
            val3.append(sheet_obj.cell(i,7).value)
        if cell_obj2.value == 'KNOT METHOD:':
            # print('KNOT METHOD: found-'+str(i))
            field.append('KNOT METHOD:')
            val1.append(sheet_obj.cell(i,7).value)
            val2.append(sheet_obj.cell(i,7).value)
            val3.append(sheet_obj.cell(i,7).value)
    
        
        cell_obj3 = sheet_obj.cell(row = 4, column = 6)
        # print(cell_obj3.value)
        quality.append(cell_obj3.value)
    
        cell_obj4 = sheet_obj.cell(row = 4, column = 8)
        # print(cell_obj4.value)
        quality.append(cell_obj4.value) 
    print("-----------------")
    print("Quality: "+quality[0])
    print("Composition: "+quality[1])
    print(val1)
    print(val2)
    print(val3)
    
    # return quality[0],quality[1],field, val1,val2,val3
    return quality[0],quality[1]







wb = load_workbook(filename='user_view/PIR10.xlsx')
print('loaded')


for sheet in wb.worksheets:
    new_wb = Workbook()
    ws = new_wb.active
    for row_data in sheet.iter_rows():
        for row_cell in row_data:
            ws[row_cell.coordinate].value = row_cell.value

    new_wb.save('splits/{0}.xlsx'.format(sheet.title))



print("scrpt execution works")

print("stage-1 ok")
cnx = mysql.connector.connect(user='root',
                              password='Password@123',
                              host='localhost',
                              database='table_test',)

print("stage-2 ok")


print("Connected to:", cnx.get_server_info())

mycursor = cnx.cursor()
mycursor.execute("USE table_test")
mycursor.execute("SHOW TABLES")

for (table_name,) in mycursor:
    print(table_name)
    if table_name=="user_view_news":
        sql = "DROP TABLE user_view_news"
        mycursor.execute(sql)
        print("table deleted....updating new one")
    

mycursor.execute("CREATE TABLE user_view_news( id BIGINT NOT NULL AUTO_INCREMENT PRIMARY KEY, name VARCHAR(255),quality VARCHAR(255),composition VARCHAR(255))")    
    
 
sql = "INSERT INTO  user_view_news(name, quality, composition) VALUES (%s, %s, %s)"

# os.system("rm splits/.DS_Store")
sheet_names=os.listdir("splits/")
mycursor2 = cnx.cursor()
a=[]
b=[]
for i in range(0, len(sheet_names)):
    a,b=scrape_sheet("splits/"+sheet_names[i])
    val=(sheet_names[i],a,b)
    mycursor2.execute(sql, val)
    print("Inserted successfully-"+str(i))    
    print("--------------")

cnx.commit()

mycursor3 = cnx.cursor()
mycursor3.execute("SELECT * FROM user_view_news")

myresult = mycursor3.fetchall()
for x in myresult:
  print(x)
  
  

cnx.close()

print("sql execution works")
