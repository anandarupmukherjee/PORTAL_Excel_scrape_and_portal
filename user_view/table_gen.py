#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Jun 14 20:06:13 2021

@author: anandarupmukherjee
https://www.w3schools.com/python/python_mysql_insert.asp
"""
import mysql.connector
from openpyxl import load_workbook, Workbook
import openpyxl as px

import subprocess
import os
import sys

def scrape_sheet(fil):
    print("scrp-sheet")
    wb_obj = px.load_workbook(fil)
    sheet_obj = wb_obj.active
      
    max_col = sheet_obj.max_column
    
    field=[]
    val1=[]
    val2=[]
    val3=[]
    quality=[]  
    composition=[]
    
    
    for i in range(1, max_col + 1):
        cell_obj = sheet_obj.cell(row = i, column = 1)
        if cell_obj.value == 'FRINGE':
            # print('fringe found-'+str(i))
            field.append('Fringe')
            val1.append(sheet_obj.cell(i,2).value)
            val2.append(sheet_obj.cell(i,3).value)
            val3.append(sheet_obj.cell(i,4).value)
        if cell_obj.value == 'MEND':
            # print('mend found-'+str(i))
            field.append('Mend')
            val1.append(sheet_obj.cell(i,2).value)
            val2.append(sheet_obj.cell(i,3).value)
            val3.append(sheet_obj.cell(i,4).value)
        if cell_obj.value == 'GR. PERCH':
            # print('perch found-'+str(i))
            field.append('Greasy Perch')
            val1.append(sheet_obj.cell(i,2).value)
            val2.append(sheet_obj.cell(i,3).value)
            val3.append(sheet_obj.cell(i,4).value)
            
            
        cell_obj2 = sheet_obj.cell(row = i, column = 6)
        if cell_obj2.value == 'MACHINE No:':
            # print('MACHINE No: found-'+str(i))
            field.append('Machine No.')
            val1.append(sheet_obj.cell(i,7).value)
            val2.append(sheet_obj.cell(i,7).value)
            val3.append(sheet_obj.cell(i,7).value)
        if cell_obj2.value == 'KNOT METHOD:':
            # print('KNOT METHOD: found-'+str(i))
            field.append('KNOT METHOD:')
            val1.append(sheet_obj.cell(i,7).value)
            val2.append(sheet_obj.cell(i,7).value)
            val3.append(sheet_obj.cell(i,7).value)
    
        
        cell_obj3 = sheet_obj.cell(row = 4, column = 6)
        # print(cell_obj3.value)
        quality.append(cell_obj3.value)
    
        cell_obj4 = sheet_obj.cell(row = 4, column = 8)
        # print(cell_obj4.value)
        quality.append(cell_obj4.value) 
    print("-----------------")
    print("Quality: "+quality[0])
    print("Composition: "+quality[1])
    print(val1)
    print(val2)
    print(val3)
    
    # return quality[0],quality[1],field, val1,val2,val3
    return quality[0],quality[1],field,val1,val2,val3







wb = load_workbook(filename='user_view/PIR10.xlsx')
print('loaded')

# while os.listdir("splits/")!=0:
for sheet in wb.worksheets:
    new_wb = Workbook()
    ws = new_wb.active
    for row_data in sheet.iter_rows():
        for row_cell in row_data:
            ws[row_cell.coordinate].value = row_cell.value

    new_wb.save('splits/{0}.xlsx'.format(sheet.title))


#import mysql.connector

cnx = mysql.connector.connect(user='alex',
                              password='Password@123',
                              host='localhost',
                              database='table_test',
                              auth_plugin='mysql_native_password')



print("Connected to:", cnx.get_server_info())


mycursor = cnx.cursor()
mycursor.execute("USE table_test")
mycursor.execute("SHOW TABLES")

for (table_name,) in mycursor:
    print(table_name)
    if table_name=="user_view_news":
        sql = "DROP TABLE user_view_news"
        mycursor.execute(sql)
        print("table deleted....updating new one")
   
#mycursor.execute("CREATE TABLE scrape(name VARCHAR(255),quality VARCHAR(255),composition VARCHAR(255))")    

mycursor.execute("CREATE TABLE user_view_news( id BIGINT NOT NULL AUTO_INCREMENT PRIMARY KEY, name VARCHAR(255),quality VARCHAR(255),composition VARCHAR(255),process1a VARCHAR(255), process1b VARCHAR(255),process1c VARCHAR(255),process2a VARCHAR(255),process2b VARCHAR(255),process2c VARCHAR(255), process3a VARCHAR(255),process3b VARCHAR(255),process3c VARCHAR(255),proc_info1 VARCHAR(255), proc_info2 VARCHAR(255)  )")    
 
 
sql = "INSERT INTO  user_view_news(name, quality, composition, process1a,process1b, process1c, process2a,process2b, process2c, process3a,process3b, process3c, proc_info1, proc_info2) VALUES (%s, %s, %s, %s, %s, %s, %s, %s,%s, %s, %s, %s,%s, %s)"

# os.system("rm splits/.DS_Store")
sheet_names=os.listdir("splits/")
mycursor2 = cnx.cursor()
a=[]
b=[]
c=[]
d=[]
e=[]
f=[]
for i in range(0, len(sheet_names)):
    a,b,c,d,e,f=scrape_sheet("splits/"+sheet_names[i])
    print(a,b)
    print(c[0],d[0],e[0],f[0])
    print(c[1],d[1],e[1],f[1])
    print(c[3],d[3], e[3],f[3])
    print(c[2],d[2])
    print(c[4],d[4])
    print("x-x-x-x-x-x-x-x-x-x")
    val=(sheet_names[i],a,b,d[0],e[0],f[0],d[1],e[1],f[1],d[3], e[3],f[3],d[2],d[4])
    mycursor2.execute(sql, val)
    print("Inserted successfully-"+str(i))    
    print("--------------")

cnx.commit()

mycursor3 = cnx.cursor()
mycursor3.execute("SELECT * FROM user_view_news")

myresult = mycursor3.fetchall()
for x in myresult:
  print(x)
  
  

cnx.close()
